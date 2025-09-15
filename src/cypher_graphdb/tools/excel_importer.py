"""Excel file importer using openpyxl row streaming (no pandas).

Each worksheet is processed via an ExcelRowSource and passed to TabularImporter.
Sheets starting with '_' treated as edges (handled by TabularImporter based on columns).
"""

import openpyxl

from cypher_graphdb import CypherGraphDB

from .excel_row_source import ExcelRowSource
from .file_importer import FileImporter
from .tabular_importer import TabularImporter


class ExcelImporter(FileImporter):
    """Excel file importer for graph data (openpyxl streaming)."""

    def __init__(self, db: CypherGraphDB):
        """Initialize Excel importer.

        Args:
            db: CypherGraphDB instance to import data into.

        """
        super().__init__(db)
        self.valid_file_extensions = (
            ".xlsx",
            ".xls",
        )

    def load_from_file(self, filename: str):
        """Load graph data from an Excel file (streamed sheet by sheet)."""
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        sheetnames = sorted(wb.sheetnames, key=lambda x: ("_" + x) if x.startswith("_") else x)
        importer = TabularImporter(self.db)
        for sheetname in sheetnames:
            label = sheetname[1:] if sheetname.startswith("_") else sheetname
            ws = wb[sheetname]
            source = ExcelRowSource(ws)
            try:
                # Get initial counts to calculate difference after import
                initial_nodes = importer.nodes_created
                initial_edges = importer.edges_created

                # Import the data
                importer.load(source, label)

                # Calculate how many objects were imported in this sheet
                nodes_imported = importer.nodes_created - initial_nodes
                edges_imported = importer.edges_created - initial_edges
                total_imported = nodes_imported + edges_imported

                # Show statistics in the import message
                self.on_import_file(filename, f"[{sheetname}]", total_imported)
            finally:
                # ExcelRowSource does not hold external resources needing close.
                pass
        wb.close()
