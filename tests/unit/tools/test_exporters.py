import csv
import os
from tempfile import TemporaryDirectory

import openpyxl

from cypher_graphdb.models import Graph, GraphEdge, GraphNode
from cypher_graphdb.tools import CsvExporter, ExcelExporter, FileExporterOptions

from ..mock_backend import build_db


def _build_sample_graph(db):
    # Create two nodes
    n1 = GraphNode(label_="Person", properties_={"gid_": "p1", "name": "Alice"})
    n2 = GraphNode(label_="Person", properties_={"gid_": "p2", "name": "Bob"})
    db.create_or_merge(n1)
    db.create_or_merge(n2)
    # Edge
    e = GraphEdge(label_="KNOWS", start_id_=n1.id_, end_id_=n2.id_, properties_={"since": 2020})
    db.create_or_merge(e)
    g = Graph()
    g.merge([n1, n2, e])
    return g


def test_csv_exporter(tmp_path):
    db = build_db()
    g = _build_sample_graph(db)
    with TemporaryDirectory(dir=tmp_path) as td:
        outdir = os.path.join(td, "csvout")
        os.makedirs(outdir, exist_ok=True)
        exporter = CsvExporter(db, FileExporterOptions(with_label=True))
        exporter.export(g, outdir)

        person_file = os.path.join(outdir, "Person.csv")
        edge_file = os.path.join(outdir, "_KNOWS.csv")
        assert os.path.exists(person_file)
        assert os.path.exists(edge_file)

        with open(person_file) as f:
            rows = list(csv.DictReader(f))
        assert {r["gid_"] for r in rows} == {"p1", "p2"}
        assert {r["name"] for r in rows} == {"Alice", "Bob"}

        with open(edge_file) as f:
            edge_rows = list(csv.DictReader(f))
        assert len(edge_rows) == 1
        er = edge_rows[0]
        assert er["start_gid_"] == "p1" and er["end_gid_"] == "p2"
        assert er["since"] == "2020"
    # td automatically cleaned


def test_excel_exporter(tmp_path):
    db = build_db()
    g = _build_sample_graph(db)
    with TemporaryDirectory(dir=tmp_path) as td:
        outdir = os.path.join(td, "xlout")
        os.makedirs(outdir, exist_ok=True)
        exporter = ExcelExporter(db, FileExporterOptions(with_label=False))
        exporter.export(g, outdir)

        person_file = os.path.join(outdir, "Person.xlsx")
        edge_file = os.path.join(outdir, "_KNOWS.xlsx")
        assert os.path.exists(person_file)
        assert os.path.exists(edge_file)

        wb = openpyxl.load_workbook(person_file, read_only=True)
        ws = wb["Person"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "label_" not in headers
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        gid_idx = headers.index("gid_")
        names = {r[headers.index("name")] for r in rows}
        gids = {r[gid_idx] for r in rows}
        assert names == {"Alice", "Bob"}
        assert gids == {"p1", "p2"}

        wb2 = openpyxl.load_workbook(edge_file, read_only=True)
        ws2 = wb2["_KNOWS"]
        headers2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
        assert "label_" not in headers2
        data_rows = list(ws2.iter_rows(min_row=2, values_only=True))
        wb2.close()
        assert len(data_rows) == 1
        dr = data_rows[0]
        hmap = {h: i for i, h in enumerate(headers2)}
        assert dr[hmap["start_gid_"]] == "p1"
        assert dr[hmap["end_gid_"]] == "p2"
        assert dr[hmap["since"]] == 2020
    # td automatically cleaned
